
import openpyxl as opx
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from collections import defaultdict
import pprint as p
import datetime as dt
import pandas as pd

wb = opx.Workbook()
# reference workbook, where data is pulled
refWB = opx.load_workbook(r"C:\Users\VD102541\Desktop\VictorChamberUsage.xlsx", data_only=True)
# indices of the various chamber-specific sheets
refSheets = refWB.worksheets[0:14]
activeSheet = wb.active
activeSheet.title = "Dashboard"


#color pallete templates, for changing cell colors
darkRedColor = PatternFill(patternType='solid', start_color='DE3163', end_color='DE3163')
darkOrangeColor = PatternFill(patternType='solid', start_color='F6C010', end_color='FF5833')
lightOrangeColor = PatternFill(patternType='solid', start_color='FFE262', end_color='FFE262')
greenColor = PatternFill(patternType='solid', start_color='43B854', end_color='43B854')
lightGreenColor = PatternFill(patternType='solid', start_color='76FF91', end_color='76FF91')
blueColor = PatternFill(patternType='solid', start_color='5A63FF', end_color='5A63FF')
lightBlueColor = PatternFill(patternType='solid', start_color='AEC1FF', end_color='AEC1FF')


#Cell with Today's Date
activeSheet['A1'] = "Today's date is "
activeSheet['A1'].alignment = Alignment(wrap_text=True)
activeSheet['B1'] = "=NOW()"
activeSheet['B1'].number_format = "MM/DD/YYYY"

#column widths
activeSheet.column_dimensions['B'].width = 15
activeSheet.column_dimensions['C'].width = 25
activeSheet.column_dimensions['D'].width = 20
activeSheet.column_dimensions['F'].width = 25


#Title cell
currentProgress = activeSheet['C3']
currentProgress.value = "Lots currently in progress: "
currentProgress.fill = darkOrangeColor
for cell in activeSheet.iter_rows(min_row=4, max_row=17, min_col=3, max_col=3):
    cell[0].fill = lightOrangeColor

#Total number of units in chambers
totalUnits = activeSheet['D3']
totalUnits.value = "Total units in progress"
totalUnits.fill = blueColor
for cell in activeSheet.iter_rows(min_row=4, max_row=17, min_col=4, max_col=4):
    cell[0].fill = lightBlueColor

#Lot Progress
lotProgress = activeSheet['F3']
lotProgress.value = "Lot Progress"
lotProgress.fill = darkRedColor
lotProgress.font = Font(color="FDFEFE")
lotProgressRefs = [f"F{i}" for i in range(4, 18)]
for index, cell in enumerate(lotProgressRefs):
    activeSheet[f'{cell}'] = refSheets[index].title
    activeSheet[f'{cell}'].alignment = Alignment(wrap_text=True)


#Chamber Cells
chamberCellRefs = [(f"B{i}") for i in range(4, 18)]
for index, cell in enumerate(chamberCellRefs):
    activeSheet[f'{cell}'] = refSheets[index].title
    activeSheet[f'{cell}'].fill = greenColor
    activeSheet[f'{cell}'].alignment = Alignment(wrap_text=True)

#First Progress Chart - num of lots, total num of units for each chamber
lots125CBake = refWB['Dashboard']['C3'].value
activeSheet['C4'].value = lots125CBake


#EXPERIMENTAL CODE

def getData(sheetIndex, columnIndex):

    """Takes the integer index of the sheet (from the list refSheets) as the first argument, the column index (as a letter) for the
       second argument"""

    dataList = list()
    for row in range(3, refSheets[sheetIndex].max_row):
        for column in columnIndex:
            cellname = f'{column}{row}'
            cellValue = refSheets[sheetIndex][f'{cellname}'].value
            if cellValue is not None:
                dataList.append((cellValue))
    return(dataList)

chamberDict = defaultdict(list) #initiates the dictionary which will hold all of our dynamic dashboard data. The keys
                                #are chambers, the values will be lists containing data from columns B(lot number),
                                #M(removal date/time), N(time until removal) and O(percentage time remaining)

for sheet in refSheets:    #Establishes chambers (i.e. titles of the worksheets) as dictionary keys
    chamberDict[f'{sheet.title}']

for index, sheet in enumerate(refSheets): #Gathers the lists of and puts it into the dictionary
    chamberDict[f'{sheet.title}'].append(getData(index, "B")) #Lot Nums (i.e. RA numbers)
    chamberDict[f'{sheet.title}'].append(getData(index, "L")) #Date/time in
    chamberDict[f'{sheet.title}'].append(getData(index, "M")) #Removal date
    chamberDict[f'{sheet.title}'].append(getData(index, "N")) #Time until removal
    chamberDict[f'{sheet.title}'].append(getData(index, "O")) # % of time remaining


for sheet in refSheets: #converts the % of time remaining into a percentage, rounding off the first decimal place
    (chamberDict[f'{sheet.title}'][-1]) = [ f'{(round(x * 100, 1))}%' for x in (chamberDict[f'{sheet.title}'][-1])]


    
#columns for the individual dataframes    
columnList = ["RA Number", "Date/Time in", "Removal Date/Time","Time Until Removal", r"% of time remaining"]


def makeDataFrame(chamberSelection):
    """Makes a data frame for each chamber"""
    df = pd.DataFrame()
    df['RA Number'] = pd.Series(chamberDict[f'{chamberSelection}'][0])
    df['Date/Time in'] = pd.Series(chamberDict[f'{chamberSelection}'][1])
    df['Removal Date/Time'] = pd.Series(chamberDict[f'{chamberSelection}'][2])
    df['Time Until Removal'] = pd.Series(chamberDict[f'{chamberSelection}'][3])
    df[r'% remaining'] = pd.Series(chamberDict[f'{chamberSelection}'][4])
    return(df)


# Each chambers' dataframe. Maybe there is a more elegant way to do this, for now it works. 

bake125CDF = makeDataFrame('125C Bake')
bake150CDF = makeDataFrame('150C Bake')
bake180CDF = makeDataFrame('180C Bake')
bake210CDF = makeDataFrame('210C Bake')
soak3060DF = makeDataFrame('30.60 Soak')
soak6060DF = makeDataFrame('85.85 Soak')
coldOven8DF = makeDataFrame('Oven 8 COLD')
uHast1DF = makeDataFrame('UHAST 1')
uHast2DF = makeDataFrame('UHAST 2')
uHast5DF = makeDataFrame('UHAST 5')
TC2DF = makeDataFrame('TC.2_D')
TC3DF = makeDataFrame('TC.3_D')
TC4DF = makeDataFrame('TC.4_D')

# A list of all the dataframes
dfList = [bake125CDF, bake150CDF, bake180CDF, 
          bake210CDF, soak3060DF, soak6060DF,
          coldOven8DF, uHast1DF, uHast2DF,
          uHast5DF, TC2DF, TC3DF, TC4DF]


wb.save(r"C:\Users\VD102541\Desktop\DashboardTest.xlsx")

#I've got solid individual data frames for each chamber. 
#Still left to do is
        # Figure out how to enter each dataframe into a corresponding row/column in the dashboard sheet.
        # Paste this code into the main updater, add something like a "view dashboard" option in the opening screen. You'll
        # be able to pick which chamber you want to view the dataframe of, or view all df's at the same time. 
