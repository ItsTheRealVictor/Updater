import openpyxl
import pandas as pd
import datetime as dt

wb = openpyxl.load_workbook(r"C:\Users\VD102541\Desktop\VictorChamberUsageTestAsset.xlsx")
# # List of sheets
sheetsList=wb.sheetnames
chambersOnlySheetsList=sheetsList[0:14]
allSheets = wb.worksheets[0:14]

for sheet in allSheets:
    sheet.cell(row=1, column=11).value = "=NOW()"

wb.save(r"C:\Users\VD102541\Desktop\VictorChamberUsageTestAsset.xlsx")