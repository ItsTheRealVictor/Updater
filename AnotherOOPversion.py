import openpyxl as opx
import pprint
import datetime
from dateutil import parser
from dateutil import relativedelta
import sys
from prompt_toolkit.layout.dimension import D
import questionary as q
import pandas as pd

wb = opx.load_workbook(r"C:\Users\VD102541\Desktop\VictorChamberUsage.xlsx")
# # List of sheets
sheetsList=wb.sheetnames
chambersOnlySheetsList=sheetsList[0:14]



class GetDataFromUser:

    def __init__(self, dataList):
        self.dataList = dataList

    dataList = []

    def getVariables():
        userLotNumber = input('Enter the Lot number: ')
        userPartNumber = input('Enter the part number: ')
        userNumOfLots = int(input('Enter the number of lots: '))
        userQuantity = int(input('Enter the quantity: '))
        userStartTime = input('Enter the starting time (HH:MM): ')
        userLotOwner = input('Enter the owner: ')
        userDateInput = input("Enter your start date(YYYY-MM-DD) or enter 't' to use today's date: ")
        if userDateInput == 't':
            userDateInput = (str(datetime.date.today()))
        userTimeIntervalInput = input('enter your time interval: ')

        varList = [userLotNumber, #(project) 
                   userPartNumber, #(product)
                   userNumOfLots, #(# of lots)
                   userQuantity, #(qty)
                   userDateInput, #(date in)
                   userStartTime, 
                   userLotOwner,
                   userTimeIntervalInput]

        for var in varList:
            GetDataFromUser.dataList.append(var)

    def storeData():
        """Stores the data input by the user into a list (dataList)"""
        for index, variable in enumerate(GetDataFromUser.dataList):
            GetDataFromUser.dataList[index] = variable

    def computeDate(incomingDate):
        """Takes the date and time interval from the user and computes the future date and day. """
        userTimeIntervalDelta = relativedelta.relativedelta(
            hours=int(GetDataFromUser.dataList[-1])) # I don't know how to access the userTimeIntervalInput variable from the local
                                                     # scope of the function getVariables(). My workaround is to access it by slicing
                                                     # out of the list I created in the for loop at the end of getVariables()
        parsedUserDateInput = parser.parse(GetDataFromUser.dataList[4]).date() # Same as above, the item at dataList[4] is the user's input date. 
        future = parsedUserDateInput + userTimeIntervalDelta #takes the user's date and time interval, adds them together
        parsedFutureDate = datetime.datetime.strftime(future, "%Y-%m-%d")#converts the future date into datetime format
        parsedUserDateInputDay = datetime.datetime.strftime(
            parsedUserDateInput, "%A")
        parsedFutureDay = datetime.datetime.strftime(future, "%A") #takes the future date from parsedFutureDate and computes the day 
                                                                   #of the week
        GetDataFromUser.dataList.insert(5, f"{parsedFutureDate}") #adds the future date and day of the week to the dataList
        return f"{parsedFutureDate}"

    def doAllFunctions():
        GetDataFromUser.getVariables()
        GetDataFromUser.storeData()
        GetDataFromUser.computeDate(GetDataFromUser.dataList[4])


class updateSpreadSheet:

    def __init__(self, chamberChoice, chamberList):
        self.chamberChoice = chamberChoice
        self.chamberList = chamberList

    #chamberChoice is an empty variable because the choice has yet to made by the user
    chamberChoice = ''
        
    chamberList = [
        "125C Bake",
        "150C Bake",
        "180C Bake",
        "210C Bake",
        "30.60 Soak",
        "60.60 Soak",
        "85.85 Soak",
        "Oven 8 COLD",
        "UHAST 1",
        "UHAST 2",
        "UHAST 5",
        "TC.2_D",
        "TC.3_D",
        "TC.4_D"
                        ]

    def pickChamber():

        chamberDict = {k: v for k, v in enumerate(updateSpreadSheet.chamberList)} # I used enumerate in case I want to add or subtract chambers in the future.
        pprint.pprint(chamberDict)
        userChamberSelection = int(input("Enter the number which corresponds to the chamber: "))
        for k, v in chamberDict.items():
            if userChamberSelection == k:
                updateSpreadSheet.chamberChoice = f"{v}"



        progBarFormulaCombinedDateTimeIn = r"""=TEXT(F3,"m/dd/yy ")&TEXT(H3,"HH:MM")""" #Combines 2 cells: Date in and Time in
        progBarFormulaCombinedDateTimeOut = r"""=TEXT(G3,"m/dd/yy ")&TEXT(H3,"HH:MM")""" #Combines Date out and Time in
        progBarFormulaRemainingTime = r"""=M3 - $O$1""" #Subtracts Current date from Date out
        progBarFormulaRemainingTimePercentage = r"""=(M3-$O$1)/(M3-L3)""" #Divides Remaining Time by Total Time

        inputDataList = [i for i in GetDataFromUser.dataList]
        inputDataList.insert(0, "") #First 2 rows are empty
        inputDataList.insert(1, "") #first 2 rows are empty
        insertLocation = 3
        for index, value in enumerate(chambersOnlySheetsList):
            if value == updateSpreadSheet.chamberChoice:
                wb[sheetsList[index]].insert_rows(insertLocation)
                for i in range(2, 10):
                    cellref = wb[sheetsList[index]].cell(row=3, column=i)
                    cellref.value = inputDataList[i]
                for j in range(12,13):
                    progBarInsertOne = wb[sheetsList[index]].cell(row=3, column=j)
                    progBarInsertOne.value = progBarFormulaCombinedDateTimeIn
                for k in range(13,14):
                    progBarInsertTwo = wb[sheetsList[index]].cell(row=3, column=k)
                    progBarInsertTwo.value = progBarFormulaCombinedDateTimeOut
                for g in range(14,15):
                    progBarInsertThree = wb[sheetsList[index]].cell(row=3, column=g)
                    progBarInsertThree.value = progBarFormulaRemainingTime
                    progBarInsertThree.number_format = "DD.HH.MM."
                for l in range(15,16):
                    progBarInsertFour = wb[sheetsList[index]].cell(row=3, column=l)
                    progBarInsertFour.value = progBarFormulaRemainingTimePercentage
                    progBarInsertFour.style = "Percent"
        #Need to add a function or otherwise better way (i.e. non-code repeating) progress dashboard functionality.
        #This works for now but needs to be changed later



class confirmUpdateSpreadsheet:

    def confirmUpdate():
        CHOICES = ['Update my spreadsheet and quit', 'Update my spreadsheet and add more parts', 'Quit']
        while True:
            confirm = q.select("What would you like to do now?", choices = CHOICES).ask()
            if confirm == 'Update my spreadsheet and quit':
                wb.save(r"C:\Users\VD102541\Desktop\VictorChamberUsage.xlsx")
                sys.exit()
            elif confirm == 'Update my spreadsheet and add more parts':
                wb.save(r"C:\Users\VD102541\Desktop\VictorChamberUsage.xlsx")
                GetDataFromUser.dataList.clear()
                GetDataFromUser.doAllFunctions()
                updateSpreadSheet.pickChamber()
                keepGoing = q.select("Do you want to keep going?", choices = ['Yes, keep going', 'Quit']).ask()
                if keepGoing == 'Yes, keep going':
                    wb.save(r"C:\Users\VD102541\Desktop\VictorChamberUsage.xlsx")
                    GetDataFromUser.dataList.clear()
                    GetDataFromUser.doAllFunctions()
                    updateSpreadSheet.pickChamber()
                elif keepGoing == 'Quit':
                    sys.exit()
            elif confirm == 'Quit':
                print("Bruh")
                sys.exit()

class OpeningScreen:

    def opening():
        while True:
            openingScreen = input("""
                #############################################################
                                    UPDATER
                                    Version 1.0
                                    by Victor Delgado
                                    ---------
                                    Press 'u' to use the updater
                                    Press 'd' to use the date calculator
                                    Press 'q' to quit
                ##############################################################
        """)
            if openingScreen == 'u':
                break
            elif openingScreen == 'd':
                startDate = input("Enter your start date (YYYY-MM-DD) or enter 't' to use today's date: ")
                if startDate == 't':
                        startDate = str(datetime.date.today())
                userTime = input("Enter your time interval (in hours): ")

                """Takes the date and time interval from the user and computes the future date and day. """
                userinputTime = relativedelta.relativedelta(hours=int(userTime)) 
                parsedUserDateInput = parser.parse(startDate).date()  
                future = parsedUserDateInput + userinputTime 
                parsedFutureDate = datetime.datetime.strftime(future, "%Y-%m-%d")#converts the future date into datetime format
                parsedUserDateInputDay = datetime.datetime.strftime(parsedUserDateInput, "%A")
                parsedFutureDay = datetime.datetime.strftime(future, "%A") #takes the future date from parsedFutureDate and computes the day 
                                                                   #of the week
                    # GetDataFromUser.dataList.insert(5, f"{parsedFutureDate} ({parsedFutureDay})") #adds the future date and day of the week to the dataList
                print(f"{userTime} hours after {parsedUserDateInputDay} {startDate} is {parsedFutureDay} {parsedFutureDate}")
            elif openingScreen == 'q':
                sys.exit()
            else:
                print("Invalid input. Try Again")

                
OpeningScreen.opening()
GetDataFromUser.doAllFunctions()
updateSpreadSheet.pickChamber()
confirmUpdateSpreadsheet.confirmUpdate()
