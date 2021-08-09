from dashboardSheet import makeDataFrame
import openpyxl as opx
import pprint
import datetime
from dateutil import parser
from dateutil import relativedelta
import sys
import questionary as q
import pandas as pd
from collections import defaultdict

wb = opx.load_workbook(r"C:\Users\valex\Desktop\VictorChamberUsageDummyData.xlsx", data_only=True)
# # List of sheets
sheetsList=wb.sheetnames # a list of the sheet names
chambersOnlySheetsList=sheetsList[0:14]
refSheets = wb.worksheets



class GetDataFromUser:

    def __init__(self, dataList):
        self.dataList = dataList

    dataList = []

    def getVariables():
        userLotNumber = input('Enter the Lot number: ')
        userPartNumber = input('Enter the part number: ')
        userNumOfLots = input('Enter the number of lots: ')
        userQuantity = input('Enter the quantity: ')
        userStartTime = input('Enter the starting time (HH:MM): ')
        userLotOwner = input('Enter the owner: ')
        userDateInput = input("Enter your start date(YYYY-MM-DD) or enter 't' to use today's date: ")
        if userDateInput == 't':
            userDateInput = (str(datetime.date.today()))
        userTimeIntervalInput = input('enter your time interval: ')

        varList = [userLotNumber, #(project)
                   userPartNumber, #(product)
                   userNumOfLots, #(# of lots)
                   userQuantity, #(qty)
                   userDateInput, #(date in)
                   userStartTime,
                   userLotOwner,
                   userTimeIntervalInput]

        for var in varList:
            GetDataFromUser.dataList.append(var)

    def storeData():
        """Stores the data input by the user into a list (dataList)"""
        for index, variable in enumerate(GetDataFromUser.dataList):
            GetDataFromUser.dataList[index] = variable

    def computeDate(incomingDate):
        """Takes the date and time interval from the user and computes the future date and day. """
        userTimeIntervalDelta = relativedelta.relativedelta(
            hours=int(GetDataFromUser.dataList[-1])) # I don't know how to access the userTimeIntervalInput variable from the local
                                                     # scope of the function getVariables(). My workaround is to access it by slicing
                                                     # out of the list I created in the for loop at the end of getVariables()
        parsedUserDateInput = parser.parse(GetDataFromUser.dataList[4]).date() # Same as above, the item at dataList[4] is the user's input date.
        future = parsedUserDateInput + userTimeIntervalDelta #takes the user's date and time interval, adds them together
        parsedFutureDate = datetime.datetime.strftime(future, "%Y-%m-%d")#converts the future date into datetime format
        parsedUserDateInputDay = datetime.datetime.strftime(
            parsedUserDateInput, "%A")
        parsedFutureDay = datetime.datetime.strftime(future, "%A") #takes the future date from parsedFutureDate and computes the day
                                                                   #of the week
        GetDataFromUser.dataList.insert(5, f"{parsedFutureDate}") #adds the future date and day of the week to the dataList
        return f"{parsedFutureDate}"

    def doAllFunctions():
        GetDataFromUser.getVariables()
        GetDataFromUser.storeData()
        GetDataFromUser.computeDate(GetDataFromUser.dataList[4])


class updateSpreadSheet:

    def __init__(self, chamberChoice, chamberList):
        self.chamberChoice = chamberChoice
        self.chamberList = chamberList

    #chamberChoice is an empty variable because the choice has yet to made by the user
    chamberChoice = ''

    chamberList = [
        "125C Bake",
        "150C Bake",
        "180C Bake",
        "210C Bake",
        "30.60 Soak",
        "60.60 Soak",
        "85.85 Soak",
        "Oven 8 Cold",
        "UHAST 1",
        "UHAST 2",
        "UHAST 5",
        "TC.2_D",
        "TC.3_D",
        "TC.4_D"
                        ]

    def pickChamber():

        chamberDict = {k: v for k, v in enumerate(updateSpreadSheet.chamberList)} # I used enumerate in case I want to add or subtract chambers in the future.
        pprint.pprint(chamberDict)
        userChamberSelection = int(input("Enter the number which corresponds to the chamber: "))
        for k, v in chamberDict.items():
            if userChamberSelection == k:
                updateSpreadSheet.chamberChoice = f"{v}"



        progBarFormulaCombinedDateTimeIn = r"""=TEXT(F3,"m/dd/yy ")&TEXT(H3,"HH:MM")"""
        progBarFormulaCombinedDateTimeOut = r"""=TEXT(G3,"m/dd/yy ")&TEXT(H3,"HH:MM")"""
        progBarFormulaRemainingTime = r"""=M3 - $O$1"""
        progBarFormulaRemainingTimePercentage = r"""=(M3-$O$1)/(M3-L3)"""
        inputDataList = [i for i in GetDataFromUser.dataList]
        inputDataList.insert(0, "")
        inputDataList.insert(1, "")
        insertLocation = 3
        for index, value in enumerate(chambersOnlySheetsList):
            if value == updateSpreadSheet.chamberChoice:
                wb[sheetsList[index]].insert_rows(insertLocation)
                for i in range(2, 10):
                    cellref = wb[sheetsList[index]].cell(row=3, column=i)
                    cellref.value = inputDataList[i]
                for j in range(12,13):
                    progBarInsertOne = wb[sheetsList[index]].cell(row=3, column=j)
                    progBarInsertOne.value = progBarFormulaCombinedDateTimeIn
                for k in range(13,14):
                    progBarInsertTwo = wb[sheetsList[index]].cell(row=3, column=k)
                    progBarInsertTwo.value = progBarFormulaCombinedDateTimeOut
                for g in range(14,15):
                    progBarInsertThree = wb[sheetsList[index]].cell(row=3, column=g)
                    progBarInsertThree.value = progBarFormulaRemainingTime
                for l in range(15,16):
                    progBarInsertFour = wb[sheetsList[index]].cell(row=3, column=l)
                    progBarInsertFour.value = progBarFormulaRemainingTimePercentage
                    progBarInsertFour.style = "Percent"
        #Need to add a function or otherwise better way (i.e. non-code repeating) progress dashboard functionality.
        #This works for now but needs to be changed later



masterChamberDataFrame = defaultdict(list)

def getData(sheetIndex, columnIndex):
    """Takes the integer index of the sheet (from the list refSheets) as the first argument,
    the column index (as a letter) for the second argument"""

    dataList = list()
    for row in range(3, refSheets[sheetIndex].max_row):
        for column in columnIndex:
            cellname = f'{column}{row}'
            cellValue = refSheets[sheetIndex][f'{cellname}'].value
            if cellValue is not None:
                dataList.append((cellValue))
    return(dataList)


# Establishes chambers (i.e. titles of the worksheets) as dictionary keys
for sheet in refSheets:
    masterChamberDataFrame[f'{sheet.title}']

# Gathers the lists of and puts it into the dictionary
for index, sheet in enumerate(refSheets):
    masterChamberDataFrame[f'{sheet.title}'].append(getData(index, "B"))  # Lot Nums (i.e. RA numbers)
    masterChamberDataFrame[f'{sheet.title}'].append(getData(index, "L"))  # Date/time in
    masterChamberDataFrame[f'{sheet.title}'].append(getData(index, "M"))  # Removal date
    # dashboardViewer.masterChamberDataFrame[f'{sheet.title}'].append(index, "N"))  # Time until removal
    masterChamberDataFrame[f'{sheet.title}'].append(getData(index, "O"))  # % of time remaining

# I have no idea why I can't round the x value to the nearest tenth. I keep getting an error saying that the round() function doesn't support strings
# I don't know why that item in the dashboardViewer.masterChamberDataFrame is a string. I don't get this error in the dashboardSheet.py file. The code is directly copied.
# I'm going to leave it in for now until later.
for sheet in refSheets:
    (masterChamberDataFrame[f'{sheet.title}'][-1]) = [(f'{(x  * 100)} %') for x in (masterChamberDataFrame[f'{sheet.title}'][-1])]

#columns for the individual dataframes
columnList = ["RA Number", "Date/Time in", "Removal Date/Time",
            "Time Until Removal", r"% of time remaining"]


def makeDataFrame(chamberSelection):
    """Makes a data frame for each chamber"""
    df = pd.DataFrame()
    df['RA Number'] = pd.Series(masterChamberDataFrame[f'{chamberSelection}'][0], dtype=str)
    df['Date/Time in'] = pd.Series(masterChamberDataFrame[f'{chamberSelection}'][1], dtype=str)
    df['Removal Date/Time'] = pd.Series(masterChamberDataFrame[f'{chamberSelection}'][2], dtype=str)
    # df['Time Until Removal'] = pd.Series(masterChamberDataFrame[f'{chamberSelection}'][3])
    df[r'% remaining'] = pd.Series(masterChamberDataFrame[f'{chamberSelection}'][3], dtype=str)

    return(df)


# Each chambers' dataframe. Maybe there is a more elegant way to do this, for now it works.

bake125CDF = makeDataFrame('125C Bake')
bake150CDF = makeDataFrame('150C Bake')
bake180CDF = makeDataFrame('180C Bake')
bake210CDF = makeDataFrame('210C Bake')
soak3060DF = makeDataFrame('30.60 Soak')
soak6060DF = makeDataFrame('85.85 Soak')
coldOven8DF = makeDataFrame('Oven 8 COLD')
uHast1DF = makeDataFrame('UHAST 1')
uHast2DF = makeDataFrame('UHAST 2')
uHast5DF = makeDataFrame('UHAST 5')
TC2DF = makeDataFrame('TC.2_D')
TC3DF = makeDataFrame('TC.3_D')
TC4DF = makeDataFrame('TC.4_D')

# A list of all the dataframes
dfList = {'125C Bake': bake125CDF,
        '1250C Bake': bake150CDF,
        '180C Bake': bake180CDF,
        '210C Bake': bake210CDF,
        '30.60 Soak': soak3060DF,
        '60.60 Soak': soak6060DF,
        'Oven 8 COLD': coldOven8DF,
        'UHAST 1': uHast1DF,
        'UHAST 2': uHast2DF,
        'UHAST 5': uHast5DF,
        'TC.2_D': TC2DF,
        'TC.3_D': TC3DF,
        'TC.4_D': TC4DF}

class confirmUpdateSpreadsheet:

    def confirmUpdate():
        CHOICES = ['Update my spreadsheet and quit', 'Update my spreadsheet and add more parts', 'Quit']
        while True:
            confirm = q.select("What would you like to do now?", choices = CHOICES).ask()
            if confirm == 'Update my spreadsheet and quit':
                wb.save(r"C:\Users\valex\Desktop\VictorChamberUsageDummyData.xlsx")
                sys.exit()
            elif confirm == 'Update my spreadsheet and add more parts':
                wb.save(r"C:\Users\valex\Desktop\VictorChamberUsageDummyData.xlsx")
                GetDataFromUser.dataList.clear()
                GetDataFromUser.doAllFunctions()
                updateSpreadSheet.pickChamber()
                keepGoing = q.select("Do you want to keep going?", choices = ['Yes, keep going', 'Quit']).ask()
                if keepGoing == 'Yes, keep going':
                    wb.save(r"C:\Users\valex\Desktop\VictorChamberUsageDummyData.xlsx")
                elif keepGoing == 'Quit':
                    wb.save(r"C:\Users\valex\Desktop\VictorChamberUsageDummyData.xlsx")
                    sys.exit()
            elif confirm == 'Quit':
                print("Bruh")
                sys.exit()

    def openingScreen():

        while True:
            openingScreen = input("""
                #############################################################
                                    UPDATER
                                    Version 1.0
                                    by Victor Delgado
                                    ---------
                                    Press 'b' to view the dashboard
                                    Press 'u' to use the updater
                                    Press 'd' to use the date calculator
                                    Press 'q' to quit
                ##############################################################
        """)
            if openingScreen == 'u':
                break
            elif openingScreen == 'b':
                dashboardChoice = q.checkbox("Choose which chamber dashboard you want to view", choices=dfList).ask()
                for choice in dashboardChoice:
                    if choice in dfList.keys():
                        print(choice)
                        print(dfList[f'{choice}'])
                sys.exit()
            elif openingScreen == 'd':
                startDate = input("Enter your start date (YYYY-MM-DD) or enter 't' to use today's date: ")
                if startDate == 't':
                        startDate = str(datetime.date.today())
                userTime = input("Enter your time interval (in hours): ")

                """Takes the date and time interval from the user and computes the future date and day. """
                userinputTime = relativedelta.relativedelta(hours=int(userTime))
                parsedUserDateInput = parser.parse(startDate).date()
                future = parsedUserDateInput + userinputTime
                parsedFutureDate = datetime.datetime.strftime(future, "%Y-%m-%d")#converts the future date into datetime format
                parsedUserDateInputDay = datetime.datetime.strftime(parsedUserDateInput, "%A")
                parsedFutureDay = datetime.datetime.strftime(future, "%A") #takes the future date from parsedFutureDate and computes the day
                                                                   #of the week
                    # GetDataFromUser.dataList.insert(5, f"{parsedFutureDate} ({parsedFutureDay})") #adds the future date and day of the week to the dataList
                print(f"{userTime} hours after {parsedUserDateInputDay} {startDate} is {parsedFutureDay} {parsedFutureDate}")
            elif openingScreen == 'q':
                sys.exit()
            else:
                print("Invalid input. Try Again")



confirmUpdateSpreadsheet.openingScreen()
GetDataFromUser.doAllFunctions()
confirmUpdateSpreadsheet.confirmUpdate()
