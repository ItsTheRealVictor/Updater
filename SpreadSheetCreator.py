import openpyxl as opx
import pprint
import datetime
from dateutil import parser
from dateutil import relativedelta


wb = opx.load_workbook(r"C:\Users\VD102541\Desktop\Copy of West Lab Chamber Usage WW25 2021.xlsx")

# List of sheets 
sheetsList = wb.sheetnames
chambersOnlySheetsList = sheetsList[0:14]


#get lot data from the user. This will be added to the spreadsheet later
userLotNumber = input('Enter the Lot number: ')
userPartNumber = input('Enter the part number: ')
userNumOfLots = input('Enter the number of lots: ')
userQuantity = input('Enter the quantity: ')
userStartTime = input('Enter the starting time: ')
userLotOwner = input('Enter the owner: ')

#get date information from the user, parse it and calculate the future date based on the user's desired time interval (how long the lot
# will stay in the chamber)
userDateInput = input('Enter your start date: ')
parsedUserDateInput = parser.parse(userDateInput).date()
parsedUserDateInputDay = datetime.datetime.strftime(parsedUserDateInput, "%A")

userTimeIntervalInput = input('enter your time interval: ')
userTimeIntervalDelta = relativedelta.relativedelta(hours=int(userTimeIntervalInput))

def calculateFutureDate(incoming):
    future = parsedUserDateInput + userTimeIntervalDelta
    parsedFutureDate = datetime.datetime.strftime(future, "%Y-%m-%d")
    parsedFutureDay = datetime.datetime.strftime(future, "%A")
    return f"{parsedFutureDate} ({parsedFutureDay})"


#    the user will enter their desired chamber. Each chamber has its own sheet, which will be updated according the to user's
#    choice of chamber
chamberChoice = ''
chamberList = {
    "1": "125C Bake",
    "2": "150C Bake",
    "3": "180C Bake",
    "4": "210C Bake",
    "5": "30.60 Soak",
    "6": "60.60 Soak",
    "7": "85.85 Soak",
    "8": "Oven 8 Cold",
    "9": "UHAST 1",
    "10": "UHAST 2", 
    "11": "UHAST 5",
    "12": "TC.2_D",
    "13": "TC.3_D",
    "14": "TC.4_D"
}

pprint.pprint(chamberList)
userChamberSelection = input("Enter the number which corresponds to the chamber:")

for k, v in chamberList.items():
    if userChamberSelection == k:
        chamberChoice = f"{v}"

userDataList = ['', '', 
                    userLotNumber, 
                    userPartNumber, 
                    userNumOfLots, 
                    userQuantity, 
                    parsedUserDateInput, 
                    calculateFutureDate(userDateInput), 
                    userStartTime, 
                    userLotOwner]

insertLocation = 3
for index, value in enumerate(chambersOnlySheetsList):
    if value == chamberChoice:
        wb[sheetsList[index]].insert_rows(insertLocation)
        for i in range(2, 10):
            cellref = wb[sheetsList[index]].cell(row=3, column=i)
            cellref.value = userDataList[i]

# def insertData():
#     for i in range(2, 10):
#         cellref = sheetBakeHTS.cell(row=3, column=i)
#         cellref.value = userDataList[i]
wb.save('InsertRowTest.xlsx')
print('Did this work?')
print('fart')


